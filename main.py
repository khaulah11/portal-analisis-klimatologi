from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
import pandas as pd
import numpy as np
import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Optional, List

app = FastAPI(title="MetClimate Sabah Climatology API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

@app.get("/")
async def read_index():
    if os.path.exists("index.html"):
        return FileResponse("index.html")
    return {"message": "index.html tidak dijumpai"}

# ============================================================
# 1. PARSER FAIL AAWS (VEKTOR PANTAS)
# ============================================================
def parse_full_aaws_file(file_bytes: bytes, filename: str):
    engine = "xlrd" if filename.endswith(".xls") else "openpyxl"
    try:
        all_sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None, engine=engine)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca fail AAWS Excel: {str(e)}")

    stations = {}

    for sheet_name, df_raw in all_sheets.items():
        if sheet_name.lower().strip() in ["datalist", "list", "index", "sheet1_copy", "summary", "sheet1"]:
            continue

        if len(df_raw) < 5:
            continue

        station_name = sheet_name
        for idx in range(min(10, len(df_raw))):
            row_str = " ".join([str(x) for x in df_raw.iloc[idx].values])
            if "station" in row_str.lower():
                val = df_raw.iloc[idx, 1] if pd.notna(df_raw.iloc[idx, 1]) else df_raw.iloc[idx, 0]
                extracted_st = str(val).replace("Station:", "").replace("Station", "").replace(":", "").strip()
                if extracted_st:
                    station_name = extracted_st
                break

        header_row = None
        for idx in range(min(20, len(df_raw))):
            row_vals = [str(x).strip().lower() for x in df_raw.iloc[idx].values]
            if "year" in row_vals and "month" in row_vals and "day" in row_vals:
                header_row = idx
                break

        if header_row is not None:
            data_df = df_raw.iloc[header_row + 1:, 0:4].copy()
            data_df.columns = ["Year", "Month", "Day", "Rainfall"]

            data_df["Year"] = pd.to_numeric(data_df["Year"], errors="coerce")
            data_df["Month"] = pd.to_numeric(data_df["Month"], errors="coerce")
            data_df["Day"] = pd.to_numeric(data_df["Day"], errors="coerce")
            data_df = data_df.dropna(subset=["Year", "Month", "Day"])

            data_df["Year"] = data_df["Year"].astype(int)
            data_df["Month"] = data_df["Month"].astype(int)
            data_df["Day"] = data_df["Day"].astype(int)

            data_df["Rainfall"] = data_df["Rainfall"].astype(str).str.strip().str.upper()
            data_df["Rainfall"] = data_df["Rainfall"].replace({"TR": "0.1", "TRACE": "0.1", "NONE": "0.0", "NULL": "nan", "-": "nan"})
            data_df["Rainfall"] = pd.to_numeric(data_df["Rainfall"], errors="coerce")
            data_df.loc[data_df["Rainfall"] < 0, "Rainfall"] = np.nan

            years = sorted(data_df["Year"].unique().tolist())
            if years:
                stations[station_name] = {
                    "sheet": sheet_name,
                    "data": data_df,
                    "years": years
                }

    return stations

# ============================================================
# 2. PARSER FAIL KAJIKLIM KONVENSIONAL
# ============================================================
def parse_conventional_file(file_bytes: bytes, filename: str):
    engine = "xlrd" if filename.endswith(".xls") else "openpyxl"
    try:
        all_sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None, engine=engine)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca fail Konvensional: {str(e)}")

    conv_stations = {}

    for sheet_name, df in all_sheets.items():
        if sheet_name.lower().strip() in ['format', 'datalist', 'list', 'index']:
            continue

        jan_row = None
        col_offset = 0
        for r in range(min(15, len(df))):
            row_str = [str(x).upper().strip() for x in df.iloc[r].values]
            if "JAN" in row_str and "FEB" in row_str:
                jan_row = r
                col_offset = row_str.index("JAN")
                break

        if jan_row is None:
            continue

        year_col = col_offset - 1
        station_name = sheet_name

        for r in range(jan_row):
            for c in range(df.shape[1]):
                v = str(df.iloc[r, c])
                if "station:" in v.lower():
                    extracted = v.split(":", 1)[-1].strip()
                    if not extracted and c + 1 < df.shape[1]:
                        extracted = str(df.iloc[r, c+1]).strip()
                    if extracted and extracted.lower() != "nan":
                        station_name = extracted
                    break

        records_by_year = {}
        for r in range(jan_row + 1, len(df)):
            yr_val = df.iloc[r, year_col]
            try:
                yr = int(float(str(yr_val).strip()))
                if yr < 1900 or yr > 2100:
                    continue
            except:
                continue

            months_val = []
            for m_idx in range(12):
                raw_v = df.iloc[r, col_offset + m_idx]
                v_str = str(raw_v).strip().replace("..", ".").replace(",", ".")
                try:
                    val = float(v_str)
                    if val < 0:
                        val = None
                except:
                    val = None
                months_val.append(round(val, 1) if val is not None else None)

            records_by_year[yr] = months_val

        if records_by_year:
            conv_stations[station_name] = {
                "sheet_name": sheet_name,
                "years": sorted(list(records_by_year.keys())),
                "monthly_by_year": records_by_year
            }

    return conv_stations

# ============================================================
# 3. PENGIRAAN VEKTOR STATISTIK
# ============================================================
def calculate_station_payload(df: pd.DataFrame, available_years: list, max_missing: int, max_consec: int, wet_th: float, suspect_th: float, extreme_th: float):
    qc_logs_all = {}
    monthly_by_year = {}
    matrix_by_year = {}
    highest_falls_by_year = {}
    highest_dates_by_year = {}

    annual_totals = []
    annual_wet_days = []
    rx1day_list = []

    for yr in available_years:
        yr_df = df[df["Year"] == yr]
        month_sums = []
        wet_days_counts = []
        rejected_months = []
        hf_list = []
        hd_list = []

        pivot = yr_df.pivot(index='Day', columns='Month', values='Rainfall').reindex(index=range(1, 32), columns=range(1, 13))
        matrix_by_year[yr] = {
            d: [round(float(v), 1) if pd.notna(v) else None for v in pivot.loc[d].tolist()] 
            for d in range(1, 32)
        }

        for m in range(1, 13):
            col_series = pivot[m] if m in pivot else pd.Series(dtype=float)
            missing_count = col_series.isna().sum()

            is_na = col_series.isna().astype(int)
            consec_missing = is_na.groupby((~is_na.astype(bool)).cumsum()).sum().max() if not col_series.empty else 31

            if missing_count > max_missing or consec_missing > max_consec or col_series.dropna().empty:
                month_sums.append(np.nan)
                wet_days_counts.append(0)
                rejected_months.append(f"{MONTH_NAMES[m-1]} {yr}")
            else:
                month_sums.append(float(col_series.sum()))
                wet_days_counts.append(int((col_series >= wet_th).sum()))

            valid_vals = col_series.dropna()
            if not valid_vals.empty and (valid_vals > 0).any():
                max_v = valid_vals.max()
                hf_list.append(round(float(max_v), 1))
                top_days = col_series[col_series == max_v].index.tolist()
                hd_list.append(",".join([str(d) for d in top_days]))
            else:
                hf_list.append(0.0)
                hd_list.append("-")

        monthly_by_year[yr] = {"totals": month_sums, "wet_days": wet_days_counts}
        highest_falls_by_year[yr] = hf_list
        highest_dates_by_year[yr] = hd_list

        annual_totals.append(round(float(yr_df["Rainfall"].sum()), 1))
        annual_wet_days.append(int((yr_df["Rainfall"] >= wet_th).sum()))
        rx1 = yr_df["Rainfall"].max()
        rx1day_list.append(round(float(rx1), 1) if pd.notna(rx1) else 0.0)

        qc_logs_all[yr] = {"rejected_months": rejected_months}

    all_totals_matrix = np.array([v["totals"] for v in monthly_by_year.values()])
    normal_mean = np.nanmean(all_totals_matrix, axis=0)
    normal_mean = np.where(np.isnan(normal_mean), 0, normal_mean).round(1).tolist()

    x = np.arange(len(available_years))
    if len(x) > 1:
        slope, intercept = np.polyfit(x, annual_totals, 1)
        trend_line = [round(float(slope * xi + intercept), 1) for xi in x]
    else:
        trend_line = annual_totals

    suspect_list = []
    for _, r in df[df["Rainfall"] > suspect_th].iterrows():
        suspect_list.append({
            "date": f"{int(r['Year'])}-{int(r['Month']):02d}-{int(r['Day']):02d}",
            "value": float(r["Rainfall"]),
            "action": "Flagged (> Suspect Limit)"
        })

    extreme_list = []
    for _, r in df[df["Rainfall"] > extreme_th].iterrows():
        extreme_list.append({
            "date": f"{int(r['Year'])}-{int(r['Month']):02d}-{int(r['Day']):02d}",
            "value": float(r["Rainfall"]),
            "action": "Flagged (> Extreme Limit)"
        })

    return {
        "available_years": available_years,
        "normal_mean": normal_mean,
        "annual": {
            "years": available_years,
            "totals": annual_totals,
            "trend_line": trend_line,
            "wet_days": annual_wet_days,
            "rx1day": rx1day_list
        },
        "monthly_by_year": {
            yr: {
                "totals": [0 if np.isnan(v) else round(v, 1) for v in monthly_by_year[yr]["totals"]],
                "wet_days": monthly_by_year[yr]["wet_days"]
            } for yr in available_years
        },
        "matrix_by_year": matrix_by_year,
        "highest_falls_by_year": highest_falls_by_year,
        "highest_dates_by_year": highest_dates_by_year,
        "qc_logs": {
            "suspect": suspect_list,
            "extreme": extreme_list,
            "by_year": qc_logs_all
        }
    }

@app.post("/api/process-aaws")
async def process_aaws(
    files: List[UploadFile] = File(...),
    max_missing: int = Form(10),
    max_consec: int = Form(4),
    wet_th: float = Form(0.1),
    suspect_th: float = Form(150.0),
    extreme_th: float = Form(250.0)
):
    all_stations_dict = {}

    for file in files:
        contents = await file.read()
        try:
            st_dict = parse_full_aaws_file(contents, file.filename)
            all_stations_dict.update(st_dict)
        except Exception:
            continue

    if not all_stations_dict:
        raise HTTPException(status_code=400, detail="Format lajur Year, Month, Day, Rainfall tidak ditemui dalam fail yang dimuat naik.")

    stations_payload = {}
    for st_name, st_info in all_stations_dict.items():
        stations_payload[st_name] = calculate_station_payload(
            st_info["data"], st_info["years"], max_missing, max_consec, wet_th, suspect_th, extreme_th
        )

    return {
        "station_list": sorted(list(stations_payload.keys())),
        "stations_data": stations_payload
    }

@app.post("/api/process-conventional")
async def process_conventional(
    files: List[UploadFile] = File(...)
):
    all_conv_dict = {}
    for file in files:
        contents = await file.read()
        try:
            c_dict = parse_conventional_file(contents, file.filename)
            all_conv_dict.update(c_dict)
        except Exception:
            continue

    if not all_conv_dict:
        raise HTTPException(status_code=400, detail="Tiada helaian cerapan konvensional sah dijumpai.")

    return {
        "conv_station_list": sorted(list(all_conv_dict.keys())),
        "conv_data": all_conv_dict
    }

# ============================================================
# 4. EXCEL BORANG EXPORT
# ============================================================
def write_single_sheet_borang(ws, station_name: str, year: int, df_station: pd.DataFrame, wet_th: float = 0.1):
    target_df = df_station[df_station["Year"] == year]

    matrix_dict = {}
    monthly_totals = []
    rain_days = []
    highest_falls = []
    highest_dates = []

    for m in range(1, 13):
        m_df = target_df[target_df["Month"] == m]
        m_sum = m_df["Rainfall"].sum() if not m_df.empty else 0.0
        monthly_totals.append(round(float(m_sum), 1) if pd.notna(m_sum) else 0.0)
        rain_days.append(int((m_df["Rainfall"] >= wet_th).sum()) if not m_df.empty else 0)

        if not m_df.empty and m_df["Rainfall"].notna().any():
            max_v = m_df["Rainfall"].max()
            if pd.notna(max_v) and max_v > 0:
                highest_falls.append(round(float(max_v), 1))
                top_days = m_df[m_df["Rainfall"] == max_v]["Day"].tolist()
                highest_dates.append(",".join([str(d) for d in top_days]))
            else:
                highest_falls.append(0.0)
                highest_dates.append("-")
        else:
            highest_falls.append(0.0)
            highest_dates.append("-")

    for d in range(1, 32):
        matrix_dict[d] = []
        for m in range(1, 13):
            val = target_df[(target_df["Month"] == m) & (target_df["Day"] == d)]["Rainfall"].values
            if len(val) > 0 and pd.notna(val[0]):
                matrix_dict[d].append(round(float(val[0]), 1))
            else:
                matrix_dict[d].append(None)

    font_title = Font(name="Arial", size=10, bold=True)
    font_header = Font(name="Arial", size=9, bold=True)
    font_data = Font(name="Arial", size=9)
    font_bold = Font(name="Arial", size=9, bold=True)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin = Side(border_style="thin", color="000000")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:M1")
    ws["A1"] = "JABATAN METEOROLOGI MALAYSIA"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:M2")
    ws["A2"] = "DAILY RAINFALL RECORD IN MILLIMETRES"
    ws["A2"].font = font_title
    ws["A2"].alignment = align_center

    ws.merge_cells("A4:H4")
    ws["A4"] = f"STATION  :  {station_name.upper()}"
    ws["A4"].font = font_bold
    ws["A4"].alignment = align_left

    ws.merge_cells("J4:M4")
    ws["J4"] = f"YEAR :  {year}"
    ws["J4"].font = font_bold
    ws["J4"].alignment = align_right

    headers = ["DATE", "JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = grid_border
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for day in range(1, 32):
        r = 6 + day
        d_cell = ws.cell(row=r, column=1, value=day)
        d_cell.font = font_header
        d_cell.alignment = align_center
        d_cell.border = grid_border

        day_vals = matrix_dict.get(day, [None] * 12)
        for m_idx in range(12):
            val = day_vals[m_idx]
            val_display = val if (val is not None and pd.notna(val)) else ""
            c = ws.cell(row=r, column=m_idx + 2, value=val_display)
            c.font = font_data
            c.alignment = align_center
            c.border = grid_border

    c_tot = ws.cell(row=38, column=1, value="TOTAL")
    c_tot.font = font_bold
    c_tot.alignment = align_center
    c_tot.border = grid_border
    for m_idx in range(12):
        c = ws.cell(row=38, column=m_idx + 2, value=monthly_totals[m_idx])
        c.font = font_bold
        c.alignment = align_center
        c.border = grid_border

    c_nod = ws.cell(row=39, column=1, value="No.Of days")
    c_nod.font = font_bold
    c_nod.alignment = align_center
    c_nod.border = grid_border
    for m_idx in range(12):
        c = ws.cell(row=39, column=m_idx + 2, value=rain_days[m_idx])
        c.font = font_bold
        c.alignment = align_center
        c.border = grid_border

    c_hf = ws.cell(row=40, column=1, value="Highest fall")
    c_hf.font = font_bold
    c_hf.alignment = align_center
    c_hf.border = grid_border
    for m_idx in range(12):
        c = ws.cell(row=40, column=m_idx + 2, value=highest_falls[m_idx])
        c.font = font_bold
        c.alignment = align_center
        c.border = grid_border

    c_dt = ws.cell(row=41, column=1, value="Date")
    c_dt.font = font_bold
    c_dt.alignment = align_center
    c_dt.border = grid_border
    for m_idx in range(12):
        c = ws.cell(row=41, column=m_idx + 2, value=highest_dates[m_idx])
        c.font = font_bold
        c.alignment = align_center
        c.border = grid_border

    ws.merge_cells("A43:D43")
    ws["A43"] = "P.K.0497(Litho)"
    ws["A43"].font = Font(name="Arial", size=8, italic=True)

    ws.merge_cells("I43:M43")
    ws["I43"] = "TR: Amount less than 0.1mm"
    ws["I43"].font = Font(name="Arial", size=8, italic=True)
    ws["I43"].alignment = align_right

    ws.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col_letter].width = 8

@app.post("/api/export-borang-excel")
async def export_borang_excel(
    files: List[UploadFile] = File(...),
    selected_station: str = Form(...),
    target_year: Optional[int] = Form(None),
    all_years: bool = Form(False),
    wet_th: float = Form(0.1)
):
    all_stations_dict = {}
    for file in files:
        contents = await file.read()
        try:
            st_dict = parse_full_aaws_file(contents, file.filename)
            all_stations_dict.update(st_dict)
        except Exception:
            continue

    if selected_station not in all_stations_dict:
        selected_station = list(all_stations_dict.keys())[0]

    st_data = all_stations_dict[selected_station]
    df = st_data["data"]
    years = st_data["years"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if all_years:
        for yr in years:
            ws = wb.create_sheet(title=str(yr))
            write_single_sheet_borang(ws, selected_station, yr, df, wet_th)
        filename_out = f"Borang_Kosong_ALL_YEARS_{selected_station.replace(' ', '_')}_{years[0]}-{years[-1]}.xlsx"
    else:
        yr = target_year if target_year in years else years[-1]
        ws = wb.create_sheet(title=str(yr))
        write_single_sheet_borang(ws, selected_station, yr, df, wet_th)
        filename_out = f"Borang_Kosong_{selected_station.replace(' ', '_')}_{yr}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename_out}"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)